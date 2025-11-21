import csv
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _check_path(path_str: str, suffix: str) -> Path:
    path = Path(path_str)
    if path.is_absolute():
        raise ValueError("Путь должен быть относительным")
    if path.suffix.lower() != suffix:
        raise ValueError("Неверный тип файла")
    return path


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
    csv_file = _check_path(csv_path, ".csv")
    xlsx_file = _check_path(xlsx_path, ".xlsx")

    if not csv_file.exists():
        raise FileNotFoundError(csv_path)

    text = csv_file.read_text(encoding="utf-8")
    if not text.strip():
        raise ValueError("Пустой CSV")

    with csv_file.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        raise ValueError("Пустой CSV")

    header = rows[0]
    if not any(header):
        raise ValueError("Нет заголовка")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    widths = []

    for row in rows:
        ws.append(row)
        for i, value in enumerate(row):
            s = "" if value is None else str(value)
            if i >= len(widths):
                widths.append(0)
            if len(s) > widths[i]:
                widths[i] = len(s)

    for i, w in enumerate(widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(w, 8)

    wb.save(str(xlsx_file))

if __name__ == "__main__":
    # Задание 3
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    with open("data/samples/people.csv", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    wb.save("data/out/people.xlsx")

    # Задание 4
    with open("data/samples/people.json", "r", encoding="utf-8") as jf:
        data = json.load(jf)

    with open("data/out/people_from_json.csv", "w", newline="", encoding="utf-8") as cf:
        fieldnames = sorted({key for obj in data for key in obj})
        writer = csv.DictWriter(cf, fieldnames=fieldnames)
        writer.writeheader()
        for row in data:
            writer.writerow(row)